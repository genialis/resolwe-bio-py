"""Annotation spreadsheet import, parsing, and validation."""

from __future__ import absolute_import, division, print_function, unicode_literals

import csv
import logging
import os
from copy import copy

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Protection
from openpyxl.worksheet.datavalidation import DataValidation

__all__ = ('FileImporter', 'FileExporter', )

logger = logging.getLogger(__name__)

# TODO: Construct these constants from the descriptor schema yaml files.

BASIC = [
    'SAMPLE_NAME',
    'FASTQ_R1',
    'FASTQ_R2',
    'COLLECTION',
    'SEQ_TYPE',
]

SAMPLE_INFO = [
    'ANNOTATOR',
    'ORGANISM',
    'SOURCE',
    'CELL_TYPE',
    'STRAIN',
    'GENOTYPE',
    'MOLECULE',
    'DESCRIPTION',
]

PROTOCOLS = [
    'GROWTH_PROTOCOL',
    'TREATMENT_PROTOCOL',
    'EXTRACT_PROTOCOL',
    'LIBRARY_PREP',
    'FRAGMENTATION_METHOD',
]

SEQ_DATA = [
    'SEQ_DATE',
    'BARCODE_REMOVED',
]

READS_DATA = [
    'BARCODE',
    'INSTRUMENT_TYPE',
    'FACILITY',
]

ANTIBODY = ['ANTIBODY']

OPTIONAL = [
    'AGE',
    'LIBRARY_STRATEGY',
    'TISSUE',
    'OTHER_CHAR_1',
    'OTHER_CHAR_2',
]

COLUMNS = (
    BASIC
    + SAMPLE_INFO
    + PROTOCOLS
    # + SEQ_DATA # TODO: Validation incompatible with Resolwe
    + READS_DATA
    + ANTIBODY
    + OPTIONAL
)

ORGANISM = {
    'Homo sapiens',
    'Mus musculus',
    'Dictyostelium discoideum',
    'Rattus norvegicus',
}

MOLECULE = {
    'total RNA',
    'polyA RNA',
    'cytoplasmic RNA',
    'nuclear RNA',
    'genomic DNA',
    'protein',
    'other',
}

SEQ_TYPE = {
    'RNA-Seq',
    'Chemical mutagenesis',
    'miRNA-Seq',
    'ncRNA-Seq'
    'RNA-Seq (CAGE)',
    'RNA-Seq (RACE)',
    'ChIP-Seq',
    'ChIPmentation',
    'ChIP-Rx',
    'MNase-Seq',
    'MBD-Seq',
    'MRE-Seq',
    'Bisulfite-Seq',
    'Bisulfite-Seq (reduced representation)',
    'MeDIP-Seq',
    'DNase-Hypersensitivity',
    'Tn-Seq',
    'FAIRE-seq',
    'SELEX',
    'RIP-Seq',
    'ChIA-PET',
    'eClIP',
    'OTHER',
}

EMPTY = {
    '',
    'N/A',
    'NONE',
    None,
}

REQUIRED = {
    'SAMPLE_NAME',
    'SEQ_TYPE',
    'ANNOTATOR',
    'ORGANISM',
    'SOURCE',
    'MOLECULE',
}

LIMITED = {
    'SEQ_TYPE': SEQ_TYPE,
    'ORGANISM': ORGANISM,
    'MOLECULE': MOLECULE,
    'BARCODE_REMOVED': {'1', '0'},
}


class FileImporter(object):
    """Import annotation spreadsheet.

    :param str annotation_path: path to a local sample annotation spreadsheet,
        or name of one in the collection
    """

    def __init__(self, annotation_path):
        """Validate the annotation sheet and create the sample list."""
        self._entry_list = []
        self.path = annotation_path
        self._is_file()
        self._populate_entries()
        self.sample_list = []
        self.invalid_samples = []
        self._create_samples()

    def _is_file(self):
        """Check is the provided path exists."""
        if not os.path.isfile(self.path):
            raise OSError(
                "The provided annotation file '{}' "
                "does not exist.".format(self.path)
            )

    def _get_spreadsheet_extension(self):
        """Find spreadsheet file extension."""
        return os.path.splitext(self.path)[1]

    def _read_xlrd(self):
        """Read Excel spreadsheet annotation file."""
        workbook = load_workbook(self.path)
        worksheet = workbook.active
        header = [cell.value for cell in worksheet[1]]
        logger.debug("Reading headers: %s", header)
        for row in worksheet.rows:
            entries = {}
            for i, cell in enumerate(row):
                if isinstance(cell.value, float):
                    entries[header[i]] = str(cell.value)
                elif cell.value in EMPTY:
                    entries[header[i]] = ''
                else:
                    entries[header[i]] = cell.value
            self._entry_list.append(entries)

    def _read_text_file(self):
        """Read simple spreadsheet annotation file."""
        with open(self.path, 'rb') as sample_sheet:
            self._entry_list = list(csv.DictReader(sample_sheet, delimiter='\t'))

    def _populate_entries(self):
        """Check the format of annotation file and assign read function."""
        if self._get_spreadsheet_extension() in ['.xls', '.xlsx', '.xlsm']:
            self._read_xlrd()
        elif self._get_spreadsheet_extension() in ['.txt', '.tab', '.tsv']:
            self._read_text_file()
        else:
            raise TypeError(
                "Annotation spreadsheet extension '{}' not recognised. Options"
                " are: '.xls', '.xlsx', '.xlsm', '.txt', '.tab', "
                "'.tsv'.".format(self._get_spreadsheet_extension())
            )

    def _create_samples(self):
        """Create a sample from each samplesheet entry."""
        for entry in self._entry_list:
            try:
                self.sample_list.append(Sample(entry))
            except ValueError as ex:
                self.invalid_samples.append(entry['SAMPLE_NAME'])
                logger.warning(ex)


class FileExporter(object):
    """Export annotation spreadsheet.

    :param str annotation_path: path to write the sample annotation spreadsheet
    :param sample_list: a list of resdk sample objects
    """

    def __init__(self, export_path=None, sample_list=[]):
        """Initialize the samplesheet template."""
        self.path = export_path
        self._samples = sample_list
        self._template = self._create_template(COLUMNS)
        for sample in sample_list:
            self._add_entry(sample, COLUMNS)

    def export_template(self):
        """Export an empty samplesheet template."""
        self._template.save(filename=self.path)

    def _create_template(self, headers):
        """Construct a template samplesheet."""
        template = Workbook()
        sheet = template.active

        # Add headers and lock the sheet
        sheet.append(headers)
        sheet.protection.sheet = True

        # Create styles
        normal = Font(name='Arial')
        bold = copy(normal)
        bold.bold = True

        # Apply formats to everything
        for cell in sheet[1]:
            header = cell.value
            col_id = cell.column
            col = sheet.column_dimensions[col_id]
            col.font = normal
            col.width = self._get_column_width(header)

            # Lock only the headers
            col.protection = Protection(locked=False)
            cell.font = normal  # Required for locking (bug?)

            # Format the required columns
            if header in REQUIRED:
                cell.font = bold

            # Format the columns with limited options
            try:
                options = '"{}"'.format(','.join(LIMITED[header]))
                valid = DataValidation(type="list", formula1=options)
                valid.error = "Invalid {}.".format(header)
                sheet.add_data_validation(valid)
                valid.add(self._get_column_body(col_id))
                col.width = self._get_column_width(LIMITED[header])
            except KeyError:
                pass

            # Format the date column
            if header == 'SEQ_DATE':
                valid_date = DataValidation(type="date")
                valid_date.error = "Invalid date."
                sheet.add_data_validation(valid_date)
                valid_date.add(self._get_column_body(col_id))

        # Return the template
        return template

    def _get_column_body(self, column):
        """Give the indices for the entire column, minus the header."""
        return '{0}2:{0}1048576'.format(column)

    def _get_column_width(self, words, factor=1.7, limits=(8, 20)):
        """Choose a column width based on the given list of words."""
        if isinstance(words, str):
            words = [words]

        width = factor * max([len(word) for word in words])

        if width > limits[1]:
            width = limits[1]
        elif width < limits[0]:
            width = limits[0]

        return width

    def _add_entry(self, sample, headers):
        """Add a sample as an entry to the samplesheet."""
        sheet = self._template.active

        # Create a lookup dictionary for the sample's meta-data
        if sample.descriptor:
            info = sample.descriptor['sample']

            # Populate the optional characteristics
            info.update(self._extract_optional(info.pop('optional_char', [])))

        else:
            info = {}

        # Populate the raw sequencing characteristics
        try:
            reads = sample.data.filter(type='data:reads')
            info.update(self._extract_seqinfo(reads[0].descriptor))
        except IndexError:
            logger.warning("No reads found for sample '%s'.", sample.name)
        except KeyError:
            logger.warning("Sample '%s' reads not annotated.", sample.name)

        lookup = {
            header.upper(): data for header, data in info.items()
            if header.upper() in COLUMNS
        }

        # Populate the primary identifiers
        lookup['SAMPLE_NAME'] = sample.name
        # lookup['COLLECTION'] = sample.collections # NOT UNIQUE

        # Eliminate null values
        for key, value in lookup.items():
            if value in EMPTY:
                lookup[key] = ''

        # Create the spreadsheet entry from the lookup dictionary
        entry = []
        for header in COLUMNS:
            try:
                entry.append(lookup[header])
            except KeyError:
                entry.append('')
        sheet.append(entry)

    def _extract_optional(self, char_list):
        """Convert a list of optional characteristics into a dictionary.

        :param list char_list: a list of strings, representing dictionary entries
            in the format "key:value"
        """
        char_dict = {}
        for char in char_list:
            pair = char.split(':')
            char_dict[pair[0]] = pair[1]

        return char_dict

    def _extract_seqinfo(self, info):
        """Extract reads annotation info from a sample."""
        entry = {'SEQ_TYPE': info['experiment_type']}
        try:
            reads_info = info['reads_info']
        except KeyError:
            reads_info = {}
        entry.update(_dict_upper(reads_info))

        try:
            protocols = info['protocols']
        except KeyError:
            protocols = {}
        entry.update(_dict_upper(protocols))

        return entry


class Sample(object):
    """Create a Sample like object.

    :param dict entry: a dictionary containing header:data pairs generated from
        an annotation spreadsheet
    """

    # TODO: Abstract this to handle other descriptor schema types.
    def __init__(self, entry):
        """Validate the entry and construct the sample descriptor."""
        self._entry = entry
        self.validate()
        self._build_descriptors()

    def _build_descriptors(self):
        """Extract the sample meta-data."""
        self.name = self._entry['SAMPLE_NAME']
        self.collection = self._entry['COLLECTION']
        self.path = self._entry['FASTQ_R1']
        self.path2 = self._entry['FASTQ_R2']
        self.seq_type = self._entry['SEQ_TYPE']

        # Build reads descriptor
        self.reads_annotation = {'protocols': {}, 'reads_info': {}}
        for char in PROTOCOLS:
            self.reads_annotation['protocols'][char.lower()] = self._entry[char]
        antibody = {
            'antibody_information': {'manufacturer': self._entry['ANTIBODY']}
        }
        self.reads_annotation['protocols'].update(antibody)

        for char in READS_DATA:
            self.reads_annotation['reads_info'][char.lower()] = self._entry[char]

        # TODO: Fix format incompatibility between openpyxl and Resolwe
        # for char in SEQ_DATA:
        #     if self._entry[char]:
        #         self.reads_annotation['reads_info'][char.lower()] = self._entry[char]

        # Build remaining sample descriptor
        self.molecule = self._entry['MOLECULE']
        self.organism = self._entry['ORGANISM']
        self.annotator = self._entry['ANNOTATOR']
        self.source = self._entry['SOURCE']
        self.sample_annotation = {
            'sample': {
                'cell_type': self._entry['CELL_TYPE'],
                'strain': self._entry['STRAIN'],
                'genotype': self._entry['GENOTYPE'],
                'description': self._entry['DESCRIPTION'],
                'optional_char': []
            }
        }

        # Include only if they are non-empty, to not override error-checking
        if self.seq_type:
            self.reads_annotation['experiment_type'] = self.seq_type

        fields = [
            ('organism', self.organism),
            ('molecule', self.molecule),
            ('annotator', self.annotator),
            ('source', self.source),
        ]
        for label, info in fields:
            if info:
                self.sample_annotation['sample'][label] = info

        # Include optional columns
        for option in sorted(OPTIONAL):
            if self._entry[option]:
                self.sample_annotation['sample']['optional_char'].append(
                    '{0}:{1}'.format(option, self._entry[option])
                )

    def validate(self):
        """Validate the annotation spreadsheet file."""
        # Check column headers
        diff1 = set(COLUMNS) - set(self._entry.keys())
        diff2 = set(self._entry.keys()) - set(COLUMNS)
        err_head = (
            "Headers '{}' {}. You should use the headers generated by"
            " the `export_annotation` method of your collection."
        )
        if diff1:
            raise KeyError(
                err_head.format("', '".join(diff1), "are missing")
            )
        if diff2:
            raise KeyError(
                err_head.format("', '".join(diff2), "not recognized")
            )

        # Check required, restricted values
        err_req = "For the sample, '{},' '{}' is not a valid {}."

        restricted = [
            ('organism', ORGANISM),
            ('molecule', MOLECULE),
            ('seq_type', SEQ_TYPE),
        ]
        for var_name, options in restricted:
            var = self._entry[var_name.upper()]
            if var not in options:
                raise ValueError(
                    err_req.format(
                        self._entry['SAMPLE_NAME'], var, var_name.upper()
                    )
                )

        # Check required, unrestricted values
        for var_name in ['annotator', 'source']:
            var = self._entry[var_name.upper()]
            if var.upper() in EMPTY:
                raise ValueError(
                    err_req.format(
                        self._entry['SAMPLE_NAME'], var, var_name.upper()
                    )
                )

    def tag_community(self):
        """Prepare community tags."""
        seq = self.seq_type.lower()
        if 'rna' in seq:
            community = 'community:rna-seq'
        elif 'chip' in seq:
            community = 'community:chip-seq'
        elif 'chemical' in seq:
            community = 'community:dicty'
        else:
            community = None
        return community


def _dict_upper(a_dict):
    """Capitalizes the keys of a dictionary."""
    return {key.upper(): value for key, value in a_dict.items()}
